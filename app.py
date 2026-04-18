import os
import io
import gzip
import json
import re
import zipfile
import threading
import requests
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory
from werkzeug.middleware.proxy_fix import ProxyFix
from dotenv import load_dotenv
from recipe_scrapers import scrape_html
from google import genai
from authlib.integrations.flask_client import OAuth

import models

load_dotenv()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-key")

# ──────────────────────────────────────
# OAuth Providers
# ──────────────────────────────────────

oauth = OAuth(app)
oauth.register(
    name="google",
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

if os.getenv("FACEBOOK_CLIENT_ID"):
    oauth.register(
        name="facebook",
        client_id=os.getenv("FACEBOOK_CLIENT_ID"),
        client_secret=os.getenv("FACEBOOK_CLIENT_SECRET"),
        authorize_url="https://www.facebook.com/v19.0/dialog/oauth",
        access_token_url="https://graph.facebook.com/v19.0/oauth/access_token",
        client_kwargs={"scope": "email public_profile"},
        userinfo_endpoint="https://graph.facebook.com/me?fields=id,name,email,picture.width(200)",
    )


def login_required_api(f):
    """Return 401 for API calls when the user is not logged in or no longer exists in the DB."""
    @wraps(f)
    def decorated(*args, **kwargs):
        uid = session.get("user_id")
        if not uid:
            return jsonify({"error": "Authentication required"}), 401
        if not models.get_user_by_id(uid):
            session.clear()
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return decorated


def _split_instructions(text):
    """Split a block of instruction text into individual steps."""
    if not text or not text.strip():
        return []
    text = text.strip()

    # Already delimited by ||
    if "||" in text:
        steps = [s.strip() for s in text.split("||") if s.strip()]
        return [re.sub(r'^Step\s+\d+\s*:\s*', '', s) for s in steps]

    # "Step 1: ... Step 2: ..." pattern
    if re.search(r'Step\s+\d+\s*:', text):
        steps = [s.strip() for s in re.split(r'(?=Step\s+\d+\s*:)', text) if s.strip()]
        return [re.sub(r'^Step\s+\d+\s*:\s*', '', s) for s in steps]

    # Numbered steps: "1. ... 2. ..." or "1) ... 2) ..."
    if re.search(r'(?:^|\n)\s*\d+[\.\)]\s', text):
        steps = re.split(r'\n\s*(?=\d+[\.\)]\s)', text)
        steps = [re.sub(r'^\d+[\.\)]\s*', '', s).strip() for s in steps if s.strip()]
        if len(steps) > 1:
            return steps

    # Double newline separated paragraphs
    if "\n\n" in text:
        steps = [s.strip() for s in text.split("\n\n") if s.strip()]
        if len(steps) > 1:
            return steps

    # Single newline separated (if lines are substantial)
    if "\n" in text:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if len(lines) > 1 and all(len(l) > 15 for l in lines):
            return lines

    return [text]

# Initialise the database on startup
models.init_db()

# Gemini client (lazy – only used when the chat endpoint is hit)
_gemini_client = None


def _get_gemini():
    global _gemini_client
    if _gemini_client is None:
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key or api_key == "your-api-key-here":
            return None
        _gemini_client = genai.Client(api_key=api_key)
    return _gemini_client


# ──────────────────────────────────────
# PWA assets (must be served from root for proper scope)
# ──────────────────────────────────────

@app.route("/sw.js")
def service_worker():
    return send_from_directory("static", "sw.js", mimetype="application/javascript")


@app.route("/manifest.json")
def manifest():
    return send_from_directory("static", "manifest.json", mimetype="application/manifest+json")


# ──────────────────────────────────────
# Pages
# ──────────────────────────────────────

@app.route("/")
@app.route("/calendar")
@app.route("/recipes")
@app.route("/favourites")
@app.route("/shopping")
@app.route("/chat")
@app.route("/settings")
@app.route("/share")
def index():
    if not session.get("user_id"):
        return redirect("/login")
    return render_template("index.html")


@app.route("/login")
def login_page():
    if session.get("user_id"):
        return redirect("/calendar")
    return render_template("login.html")


# ──────────────────────────────────────
# Auth routes
# ──────────────────────────────────────

@app.route("/auth/google")
def auth_google():
    base_url = os.getenv("BASE_URL")
    if base_url:
        redirect_uri = base_url.rstrip("/") + "/auth/callback"
    else:
        redirect_uri = url_for("auth_callback", _external=True)
    return oauth.google.authorize_redirect(redirect_uri)


@app.route("/auth/callback")
def auth_callback():
    token = oauth.google.authorize_access_token()
    userinfo = token.get("userinfo") or oauth.google.userinfo()
    user = models.get_or_create_user(
        google_id=userinfo["sub"],
        email=userinfo.get("email", ""),
        name=userinfo.get("name"),
        picture=userinfo.get("picture"),
    )
    # Adopt orphan data (pre-auth rows) on first-ever login
    models.adopt_orphan_data(user["id"])
    session["user_id"] = user["id"]
    session["user_name"] = user.get("name", "")
    session["user_picture"] = user.get("picture", "")
    return redirect("/calendar")


@app.route("/auth/facebook")
def auth_facebook():
    base_url = os.getenv("BASE_URL")
    if base_url:
        redirect_uri = base_url.rstrip("/") + "/auth/facebook/callback"
    else:
        redirect_uri = url_for("auth_facebook_callback", _external=True)
    return oauth.facebook.authorize_redirect(redirect_uri)


@app.route("/auth/facebook/callback")
def auth_facebook_callback():
    token = oauth.facebook.authorize_access_token()
    resp = oauth.facebook.get("https://graph.facebook.com/me?fields=id,name,email,picture.width(200)")
    userinfo = resp.json()
    picture = None
    if userinfo.get("picture", {}).get("data", {}).get("url"):
        picture = userinfo["picture"]["data"]["url"]
    user = models.get_or_create_user_facebook(
        facebook_id=userinfo["id"],
        email=userinfo.get("email", ""),
        name=userinfo.get("name"),
        picture=picture,
    )
    models.adopt_orphan_data(user["id"])
    session["user_id"] = user["id"]
    session["user_name"] = user.get("name", "")
    session["user_picture"] = user.get("picture", "")
    return redirect("/calendar")


@app.route("/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return redirect("/login")


@app.route("/api/me")
def api_me():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"error": "Not authenticated"}), 401
    user = models.get_user_by_id(uid)
    if not user:
        session.clear()
        return jsonify({"error": "User not found"}), 401
    return jsonify({
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "picture": user["picture"],
    })


# ──────────────────────────────────────
# Recipe API
# ──────────────────────────────────────

@app.route("/api/recipes", methods=["GET"])
@login_required_api
def api_get_recipes():
    return jsonify(models.get_all_recipes(session["user_id"]))


@app.route("/api/recipes/<int:recipe_id>", methods=["GET"])
@login_required_api
def api_get_recipe(recipe_id):
    recipe = models.get_recipe(session["user_id"], recipe_id)
    if recipe is None:
        return jsonify({"error": "Recipe not found"}), 404
    return jsonify(recipe)


@app.route("/api/recipes", methods=["POST"])
@login_required_api
def api_create_recipe():
    data = request.get_json(force=True)
    title = data.get("title", "").strip()
    ingredients = data.get("ingredients", [])
    instructions = data.get("instructions", [])
    if not title or not ingredients:
        return jsonify({"error": "Title and ingredients are required"}), 400
    try:
        recipe_id = models.create_recipe(
            session["user_id"],
            title=title,
            ingredients=ingredients,
            instructions=instructions,
            source_url=data.get("source_url"),
            image_url=data.get("image_url"),
            total_time=data.get("total_time"),
            servings=data.get("servings"),
            categories=data.get("categories"),
        )
    except Exception as e:
        return jsonify({"error": f"Failed to save recipe: {e}"}), 500
    return jsonify({"id": recipe_id}), 201


@app.route("/api/recipes/<int:recipe_id>", methods=["DELETE"])
@login_required_api
def api_delete_recipe(recipe_id):
    models.delete_recipe(session["user_id"], recipe_id)
    return jsonify({"ok": True})


@app.route("/api/recipes/bulk-delete", methods=["POST"])
@login_required_api
def api_bulk_delete_recipes():
    data = request.get_json(force=True)
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "No recipe IDs provided"}), 400
    models.delete_recipes_bulk(session["user_id"], ids)
    return jsonify({"ok": True, "deleted": len(ids)})


@app.route("/api/recipes/<int:recipe_id>", methods=["PUT"])
@login_required_api
def api_update_recipe(recipe_id):
    data = request.get_json(force=True)
    title = data.get("title", "").strip()
    ingredients = data.get("ingredients", [])
    instructions = data.get("instructions", [])
    if not title or not ingredients:
        return jsonify({"error": "Title and ingredients are required"}), 400
    models.update_recipe(
        session["user_id"],
        recipe_id,
        title=title,
        ingredients=ingredients,
        instructions=instructions,
        source_url=data.get("source_url"),
        image_url=data.get("image_url"),
        total_time=data.get("total_time"),
        servings=data.get("servings"),
        categories=data.get("categories"),
    )
    return jsonify({"ok": True})


@app.route("/api/recipes/<int:recipe_id>/favourite", methods=["POST"])
@login_required_api
def api_toggle_favourite(recipe_id):
    result = models.toggle_favourite(session["user_id"], recipe_id)
    if result is None:
        return jsonify({"error": "Recipe not found"}), 404
    return jsonify({"is_favourite": result})


# ──────────────────────────────────────
# URL Recipe Scraper
# ──────────────────────────────────────

_SOCIAL_VIDEO_PATTERNS = re.compile(
    r"(tiktok\.com|vm\.tiktok\.com|instagram\.com/(reel|p)/|instagram\.com/.*/(reel|p)/|facebook\.com/(reel|watch)|fb\.watch)",
    re.IGNORECASE,
)


def _is_social_video_url(url):
    return bool(_SOCIAL_VIDEO_PATTERNS.search(url))


def _extract_recipe_from_social_video(url):
    """Fetch a social media video page and use Gemini to extract the recipe."""
    from bs4 import BeautifulSoup

    client = _get_gemini()
    if client is None:
        return None, "Gemini API key is required to import recipes from video URLs."

    # Fetch the page HTML
    resp = requests.get(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        },
        timeout=15,
        allow_redirects=True,
    )
    html = resp.text
    final_url = resp.url  # after redirects

    soup = BeautifulSoup(html, "html.parser")

    # Collect useful metadata from the page
    parts = []

    # og/meta tags
    for tag in soup.find_all("meta"):
        prop = tag.get("property", "") or tag.get("name", "")
        content = tag.get("content", "")
        if content and any(k in prop.lower() for k in ("title", "description", "og:")):
            parts.append(f"{prop}: {content}")

    # JSON-LD blocks
    for script in soup.find_all("script", type="application/ld+json"):
        if script.string:
            parts.append(script.string[:5000])

    # Visible page text (limited)
    body_text = soup.get_text(separator="\n", strip=True)[:8000]
    parts.append(body_text)

    page_content = "\n---\n".join(parts)

    # Extract image from og:image
    og_image = None
    og_tag = soup.find("meta", property="og:image")
    if og_tag:
        og_image = og_tag.get("content")

    prompt = (
        "The following is the text content extracted from a social media video page "
        "(TikTok or Instagram) that contains a cooking recipe video.\n\n"
        f"Page URL: {final_url}\n\n"
        f"Page content:\n{page_content}\n\n"
        "Extract the recipe from this page content. If the caption or description mentions "
        "ingredients or cooking steps, use those. If the information is incomplete, fill in "
        "reasonable details based on the recipe title/description.\n\n"
        "Return ONLY a valid JSON object (no markdown fences) with this structure:\n"
        '{"title": "...", "ingredients": ["..."], "instructions": ["step 1", "step 2", ...], '
        '"total_time": "... min", "servings": "...", "categories": ["..."]}\n\n'
        "Use METRIC measurements (grams, ml, °C). "
        "If you cannot find any recipe in the content, return: "
        '{"error": "No recipe found in this video"}'
    )

    lang_key = None
    try:
        uid = session.get("user_id")
        if uid:
            lang_key = models.get_setting(uid, "language", "en")
    except Exception:
        pass

    system = (
        "You are a recipe extraction assistant. Extract cooking recipes from social media "
        "video page content. Always use METRIC measurements. Return only valid JSON."
    )
    if lang_key == "fr":
        system += (
            " The user speaks French. Recipe title, ingredients, and instructions "
            "should be in French. JSON keys must remain in English."
        )

    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=[{"role": "user", "parts": [{"text": prompt}]}],
        config={
            "system_instruction": system,
            "temperature": 0.3,
            "max_output_tokens": 4000,
        },
    )
    raw = response.text.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

    parsed = json.loads(raw)

    if "error" in parsed:
        return None, parsed["error"]

    return {
        "title": parsed.get("title", "Untitled Recipe"),
        "ingredients": parsed.get("ingredients", []),
        "instructions": parsed.get("instructions", []),
        "image_url": og_image,
        "total_time": parsed.get("total_time"),
        "servings": parsed.get("servings"),
        "source_url": final_url,
        "categories": parsed.get("categories", []),
    }, None


@app.route("/api/scrape", methods=["POST"])
@login_required_api
def api_scrape_url():
    data = request.get_json(force=True)
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL is required"}), 400

    # Social media video URLs → use Gemini to extract recipe
    if _is_social_video_url(url):
        try:
            result, error = _extract_recipe_from_social_video(url)
            if error:
                return jsonify({"error": error}), 400
            return jsonify(result)
        except Exception as exc:
            return jsonify({"error": f"Could not extract recipe from video: {exc}"}), 400

    try:
        html = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/120.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            },
            timeout=15,
        ).text
        scraper = scrape_html(html=html, org_url=url)

        title = scraper.title() or "Untitled Recipe"
        ingredients = scraper.ingredients() or []
        instructions_raw = scraper.instructions() or ""

        # instructions() returns a single string – split into steps
        if isinstance(instructions_raw, str):
            instructions = [
                s.strip() for s in re.split(r"\n+", instructions_raw) if s.strip()
            ]
        else:
            instructions = list(instructions_raw)

        image = None
        try:
            image = scraper.image()
        except Exception:
            pass

        total_time = None
        try:
            total_time = str(scraper.total_time()) + " min"
        except Exception:
            pass

        servings = None
        try:
            servings = str(scraper.yields())
        except Exception:
            pass

        return jsonify({
            "title": title,
            "ingredients": ingredients,
            "instructions": instructions,
            "image_url": image,
            "total_time": total_time,
            "servings": servings,
            "source_url": url,
        })

    except Exception as exc:
        return jsonify({"error": f"Could not scrape recipe: {exc}"}), 400


@app.route("/api/import-url", methods=["POST"])
@login_required_api
def api_import_url():
    """One-step import: scrape a URL and save it directly. Used by iOS Shortcuts."""
    data = request.get_json(force=True)
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL is required"}), 400

    user_id = session["user_id"]
    recipe_data = None

    # Step 1: scrape / extract
    if _is_social_video_url(url):
        result, error = _extract_recipe_from_social_video(url)
        if error:
            return jsonify({"error": error}), 400
        recipe_data = result
    else:
        try:
            html = requests.get(
                url,
                headers={
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                                  "Chrome/120.0.0.0 Safari/537.36",
                },
                timeout=15,
            ).text
            scraper = scrape_html(html=html, org_url=url)
            title = scraper.title() or "Untitled Recipe"
            ingredients = scraper.ingredients() or []
            instructions_raw = scraper.instructions() or ""
            if isinstance(instructions_raw, str):
                instructions = [s.strip() for s in re.split(r"\n+", instructions_raw) if s.strip()]
            else:
                instructions = list(instructions_raw)
            image = None
            try:
                image = scraper.image()
            except Exception:
                pass
            total_time = None
            try:
                total_time = str(scraper.total_time()) + " min"
            except Exception:
                pass
            servings = None
            try:
                servings = str(scraper.yields())
            except Exception:
                pass
            recipe_data = {
                "title": title,
                "ingredients": ingredients,
                "instructions": instructions,
                "image_url": image,
                "total_time": total_time,
                "servings": servings,
                "source_url": url,
            }
        except Exception as exc:
            return jsonify({"error": f"Could not scrape recipe: {exc}"}), 400

    # Step 2: save
    title = recipe_data.get("title", "").strip()
    ingredients = recipe_data.get("ingredients", [])
    if not title or not ingredients:
        return jsonify({"error": "Could not extract a valid recipe"}), 400

    try:
        recipe_id = models.create_recipe(
            user_id,
            title=title,
            ingredients=ingredients,
            instructions=recipe_data.get("instructions", []),
            source_url=recipe_data.get("source_url"),
            image_url=recipe_data.get("image_url"),
            total_time=recipe_data.get("total_time"),
            servings=recipe_data.get("servings"),
            categories=recipe_data.get("categories"),
        )
    except Exception as e:
        return jsonify({"error": f"Failed to save recipe: {e}"}), 500

    return jsonify({"success": True, "id": recipe_id, "title": title})


# ──────────────────────────────────────
# Import Paprika
# ──────────────────────────────────────

@app.route("/api/import-paprika", methods=["POST"])
@login_required_api
def api_import_paprika():
    """Import recipes from a .paprikarecipes file."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    user_id = session["user_id"]
    imported = []
    try:
        raw = f.read()

        recipes_json = []

        # Try ZIP-inside-gzip first (.paprikarecipes)
        try:
            decompressed = gzip.decompress(raw)
            zf = zipfile.ZipFile(io.BytesIO(decompressed))
            for name in zf.namelist():
                entry_data = zf.read(name)
                try:
                    recipe_data = gzip.decompress(entry_data)
                except Exception:
                    recipe_data = entry_data
                recipes_json.append(json.loads(recipe_data))
        except Exception:
            try:
                zf = zipfile.ZipFile(io.BytesIO(raw))
                for name in zf.namelist():
                    entry_data = zf.read(name)
                    try:
                        recipe_data = gzip.decompress(entry_data)
                    except Exception:
                        recipe_data = entry_data
                    recipes_json.append(json.loads(recipe_data))
            except Exception:
                try:
                    decompressed = gzip.decompress(raw)
                    recipes_json.append(json.loads(decompressed))
                except Exception:
                    recipes_json.append(json.loads(raw))

        for r in recipes_json:
            title = r.get("name", "").strip()
            if not title:
                continue
            ingredients_raw = r.get("ingredients", "")
            ingredients = [line.strip() for line in ingredients_raw.split("\n") if line.strip()]
            directions_raw = r.get("directions", "")
            instructions = [line.strip() for line in directions_raw.split("\n") if line.strip()]
            servings = r.get("servings", None)
            total_time = r.get("total_time", None)
            source_url = r.get("source_url") or r.get("url") or None
            image_url = r.get("image_url") or r.get("photo_url") or None
            cat_raw = r.get("categories", "")
            categories = [c.strip() for c in cat_raw.split("\n") if c.strip()] if isinstance(cat_raw, str) else (cat_raw or [])

            recipe_id = models.create_recipe(
                user_id,
                title=title,
                ingredients=ingredients if ingredients else [""],
                instructions=instructions if instructions else [""],
                source_url=source_url,
                image_url=image_url,
                total_time=total_time,
                servings=servings,
                categories=categories,
            )
            imported.append({"id": recipe_id, "title": title})

    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400

    return jsonify({"imported": imported, "count": len(imported)})


@app.route("/api/import-excel", methods=["POST"])
@login_required_api
def api_import_excel():
    """Import recipes from an Excel (.xlsx) file."""
    import openpyxl

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    user_id = session["user_id"]
    imported = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True)
        # Prefer "Full Recipes" sheet if it exists, otherwise use active sheet
        ws = wb["Full Recipes"] if "Full Recipes" in wb.sheetnames else wb.active

        # Read header row to map columns
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {h.strip().lower(): i for i, h in enumerate(headers) if h}

        def col(row, *names):
            for name in names:
                if name in header_map:
                    val = row[header_map[name]]
                    return val if val is not None else ""
            return ""

        for row in ws.iter_rows(min_row=2, values_only=True):
            title = str(col(row, "recipe name", "title", "name")).strip()
            if not title:
                continue

            # Ingredients: pipe-separated or newline-separated
            ing_raw = str(col(row, "ingredients"))
            pantry_raw = str(col(row, "pantry staples", "pantry"))
            ingredients = [i.strip() for i in re.split(r'[|\n]', ing_raw) if i.strip()]
            if pantry_raw:
                pantry = [p.strip() for p in re.split(r'[|\n]', pantry_raw) if p.strip()]
                ingredients.extend(pantry)

            # Instructions: split on || or "Step N:" boundaries, then clean prefixes
            method_raw = str(col(row, "method", "instructions", "directions"))
            instructions = _split_instructions(method_raw)

            # Cook time / total time
            total_time_raw = col(row, "total time", "total_time")
            cook_time_raw = col(row, "cook time 2 servings (mins)", "cook time",
                                "cook time 4 servings (mins)")
            if total_time_raw and str(total_time_raw).strip():
                total_time = str(total_time_raw).strip()
            elif cook_time_raw and str(cook_time_raw).strip():
                ct = str(cook_time_raw).strip()
                # If it's already a human string like "10 mins", use as-is
                if any(c.isalpha() for c in ct):
                    total_time = ct
                else:
                    total_time = f"{int(ct)} min"
            else:
                total_time = None

            # Servings: default to 2 for Gousto-style
            servings = str(col(row, "servings")) or "2 servings"
            if servings and not any(c.isdigit() for c in str(servings)):
                servings = "2 servings"

            # Categories from cuisine column
            cuisine = str(col(row, "categories", "category", "cuisine")).strip()
            categories = [cuisine] if cuisine else []

            source_url = str(col(row, "source url", "source_url", "link", "url")).strip() or None

            recipe_id = models.create_recipe(
                user_id,
                title=title,
                ingredients=ingredients if ingredients else [""],
                instructions=instructions if instructions else [""],
                source_url=source_url,
                image_url=None,
                total_time=total_time,
                servings=servings,
                categories=categories,
            )
            imported.append({"id": recipe_id, "title": title, "source_url": source_url})

        wb.close()

    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel file: {str(e)}"}), 400

    # Fetch images in background so the import returns immediately
    recipes_to_scrape = [(r["id"], r.get("source_url")) for r in imported if r.get("source_url")]
    if recipes_to_scrape:
        def _fetch_images(pairs):
            import time
            for rid, url in pairs:
                try:
                    img = _scrape_image_url(url)
                    if img:
                        models.update_recipe_image(rid, img)
                    time.sleep(0.5)  # be polite to the server
                except Exception:
                    pass
        thread = threading.Thread(target=_fetch_images, args=(recipes_to_scrape,), daemon=True)
        thread.start()

    return jsonify({"imported": imported, "count": len(imported)})


@app.route("/api/categories", methods=["GET"])
@login_required_api
def api_get_categories():
    """Return all unique categories used across recipes."""
    recipes = models.get_all_recipes(session["user_id"])
    cats = set()
    for r in recipes:
        for c in r.get("categories", []):
            cats.add(c)
    return jsonify(sorted(cats))


# ──────────────────────────────────────
# Calendar API
# ──────────────────────────────────────

@app.route("/api/calendar", methods=["GET"])
@login_required_api
def api_get_calendar():
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        return jsonify({"error": "year and month are required"}), 400
    entries = models.get_entries_for_month(session["user_id"], year, month)
    return jsonify(entries)


@app.route("/api/calendar/entries", methods=["POST"])
@login_required_api
def api_add_calendar_entry():
    data = request.get_json(force=True)
    entry_date = data.get("entry_date", "").strip()
    meal_type = data.get("meal_type", "").lower()
    recipe_id = data.get("recipe_id")
    note = data.get("note")
    servings = data.get("servings", 2)
    if not entry_date or not meal_type:
        return jsonify({"error": "entry_date and meal_type are required"}), 400
    if not recipe_id and not note:
        return jsonify({"error": "Either recipe_id or note is required"}), 400
    entry_id = models.add_calendar_entry(session["user_id"], entry_date, meal_type,
                                         recipe_id=recipe_id, note=note,
                                         servings=servings)
    return jsonify({"id": entry_id}), 201


@app.route("/api/calendar/entries/<int:entry_id>", methods=["DELETE"])
@login_required_api
def api_remove_calendar_entry(entry_id):
    models.remove_calendar_entry(session["user_id"], entry_id)
    return jsonify({"ok": True})


@app.route("/api/calendar/entries/<int:entry_id>/move", methods=["PATCH"])
@login_required_api
def api_move_calendar_entry(entry_id):
    data = request.get_json(force=True)
    entry_date = data.get("entry_date", "").strip()
    meal_type = data.get("meal_type", "").lower()
    if not entry_date or not meal_type:
        return jsonify({"error": "entry_date and meal_type are required"}), 400
    models.move_calendar_entry(session["user_id"], entry_id, entry_date, meal_type)
    return jsonify({"ok": True})


@app.route("/api/calendar/entries/<int:entry_id>/servings", methods=["PATCH"])
@login_required_api
def api_update_entry_servings(entry_id):
    data = request.get_json(force=True)
    servings = data.get("servings")
    if not servings or int(servings) < 1:
        return jsonify({"error": "servings must be >= 1"}), 400
    models.update_calendar_entry_servings(session["user_id"], entry_id, int(servings))
    return jsonify({"ok": True})


@app.route("/api/calendar/entries/<int:entry_id>/note", methods=["PATCH"])
@login_required_api
def api_update_entry_note(entry_id):
    data = request.get_json(force=True)
    note = data.get("note", "").strip()
    if not note:
        return jsonify({"error": "note is required"}), 400
    models.update_calendar_entry_note(session["user_id"], entry_id, note)
    return jsonify({"ok": True})


@app.route("/api/calendar/entries/<int:entry_id>/copy", methods=["POST"])
@login_required_api
def api_copy_calendar_entry(entry_id):
    data = request.get_json(force=True)
    entry_date = data.get("entry_date", "").strip()
    meal_type = data.get("meal_type", "").lower()
    if not entry_date or not meal_type:
        return jsonify({"error": "entry_date and meal_type are required"}), 400
    new_id = models.copy_calendar_entry(session["user_id"], entry_id, entry_date, meal_type)
    if new_id is None:
        return jsonify({"error": "Entry not found"}), 404
    return jsonify({"id": new_id}), 201


@app.route("/api/categorize-ingredient", methods=["POST"])
@login_required_api
def api_categorize_ingredient():
    data = request.get_json(force=True)
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"category": "Other"})
    return jsonify({"category": models.categorize_ingredient(text)})


@app.route("/api/custom-shopping-items", methods=["GET"])
@login_required_api
def api_get_custom_shopping_items():
    week_start = request.args.get("week_start")
    return jsonify(models.get_custom_shopping_items(session["user_id"], week_start))


@app.route("/api/custom-shopping-items", methods=["POST"])
@login_required_api
def api_add_custom_shopping_item():
    data = request.get_json(force=True)
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400
    category = models.categorize_ingredient(text)
    week_start = data.get("week_start")
    item_id = models.add_custom_shopping_item(session["user_id"], text, category, week_start)
    return jsonify({"id": item_id, "text": text, "category": category}), 201


@app.route("/api/custom-shopping-items/<int:item_id>", methods=["DELETE"])
@login_required_api
def api_delete_custom_shopping_item(item_id):
    models.delete_custom_shopping_item(session["user_id"], item_id)
    return jsonify({"ok": True})


@app.route("/api/custom-shopping-items", methods=["DELETE"])
@login_required_api
def api_clear_custom_shopping_items():
    week_start = request.args.get("week_start")
    models.clear_custom_shopping_items(session["user_id"], week_start)
    return jsonify({"ok": True})


@app.route("/api/shopping-list", methods=["GET"])
@login_required_api
def api_shopping_list():
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    if not start or not end:
        return jsonify({"error": "start and end dates are required"}), 400
    items = models.get_shopping_list_for_range(session["user_id"], start, end)
    return jsonify(items)


@app.route("/api/fix-instructions", methods=["POST"])
@login_required_api
def api_fix_instructions():
    """Re-split recipes that have all instructions in a single step."""
    user_id = session["user_id"]
    conn = models.get_db()
    rows = conn.execute("SELECT id, instructions FROM recipes WHERE user_id = ?", (user_id,)).fetchall()
    fixed = 0
    for row in rows:
        instr = json.loads(row["instructions"])
        if len(instr) == 1 and len(instr[0]) > 80:
            new_steps = _split_instructions(instr[0])
            if len(new_steps) > 1:
                conn.execute("UPDATE recipes SET instructions = ? WHERE id = ?",
                             (json.dumps(new_steps), row["id"]))
                fixed += 1
    conn.commit()
    conn.close()
    return jsonify({"fixed": fixed})


@app.route("/api/generate-meal-plan", methods=["POST"])
@login_required_api
def api_generate_meal_plan():
    """Take an AI-generated meal_plan object, save all recipes, create calendar entries."""
    from datetime import datetime, timedelta
    data = request.get_json(force=True)
    entries = data.get("entries", [])
    if not entries:
        return jsonify({"error": "No entries provided"}), 400

    user_id = session["user_id"]

    # Collect unique dates from entries, or default to next 7 days
    today = datetime.now().date()
    entry_dates = sorted(set(e.get("day", "") for e in entries if e.get("day")))
    if not entry_dates:
        days_until_monday = (7 - today.weekday()) % 7 or 7
        start = today + timedelta(days=days_until_monday)
        entry_dates = [(start + timedelta(days=i)).isoformat() for i in range(7)]

    saved = []
    for entry in entries:
        recipe_data = entry.get("recipe", {})
        title = recipe_data.get("title", "").strip()
        if not title:
            continue
        ingredients = recipe_data.get("ingredients", [])
        instructions = recipe_data.get("instructions", [])
        recipe_id = models.create_recipe(
            user_id,
            title=title,
            ingredients=ingredients if ingredients else [""],
            instructions=instructions if instructions else [""],
            source_url=recipe_data.get("source_url"),
            image_url=recipe_data.get("image_url"),
            total_time=recipe_data.get("total_time"),
            servings=recipe_data.get("servings"),
            categories=recipe_data.get("categories"),
        )
        day = entry.get("day", "")
        meal = entry.get("meal", "").lower()
        servings_num = 2
        try:
            s = recipe_data.get("servings", "")
            m = re.search(r'(\d+)', str(s))
            if m:
                servings_num = int(m.group(1))
        except Exception:
            pass
        models.add_calendar_entry(user_id, day, meal, recipe_id=recipe_id, servings=servings_num)
        saved.append({"title": title, "day": day, "meal": meal})

    return jsonify({"entries_created": len(saved), "saved": saved}), 201


@app.route("/api/modify-recipe", methods=["POST"])
@login_required_api
def api_modify_recipe():
    """Ask AI to modify an existing recipe based on a user request."""
    client = _get_gemini()
    if client is None:
        return jsonify({"error": "Gemini API key not configured."}), 503

    data = request.get_json(force=True)
    recipe = data.get("recipe", {})
    modification = data.get("modification", "")

    if not recipe or not modification:
        return jsonify({"error": "Missing recipe or modification."}), 400

    prompt = (
        f"Here is a recipe:\n"
        f"Title: {recipe.get('title', '')}\n"
        f"Ingredients: {json.dumps(recipe.get('ingredients', []))}\n"
        f"Instructions: {json.dumps(recipe.get('instructions', []))}\n"
        f"Total time: {recipe.get('total_time', 'N/A')}\n"
        f"Servings: {recipe.get('servings', 'N/A')}\n\n"
        f"Please modify this recipe to make it: {modification}\n\n"
        "Keep the same general dish but adapt the ingredients and instructions as needed. "
        "Return ONLY a JSON object (no markdown) with this structure: "
        '{"recipe": {"title": "...", "ingredients": ["..."], "instructions": ["..."], '
        '"total_time": "...", "servings": "...", "categories": ["..."]}}'
    )

    system = (
        "You are a creative chef who adapts recipes. Always use METRIC measurements "
        "(grams, kilograms, millilitres, litres, Celsius). "
        "Return only valid JSON, no markdown fences."
    )
    lang = models.get_setting(session["user_id"], "language", "en")
    if lang == "fr":
        system += (
            " The user speaks French. Recipe title, ingredients, and instructions should be in French. "
            "JSON keys must remain in English."
        )

    try:
        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=[{"role": "user", "parts": [{"text": prompt}]}],
            config={
                "system_instruction": system,
                "temperature": 0.8,
                "max_output_tokens": 3000,
            },
        )
        raw = "".join(
            p.text for p in response.candidates[0].content.parts
            if hasattr(p, "text") and p.text
        ).strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        parsed = json.loads(raw)
        return jsonify(parsed)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/regenerate-recipe", methods=["POST"])
@login_required_api
def api_regenerate_recipe():
    """Ask AI to generate one replacement recipe for a specific meal slot."""
    client = _get_gemini()
    if client is None:
        return jsonify({"error": "Gemini API key not configured."}), 503

    data = request.get_json(force=True)
    day = data.get("day", "")
    meal = data.get("meal", "")
    old_title = data.get("old_title", "")
    context = data.get("context", "")

    prompt = (
        f"I need a replacement recipe for {meal} on {day}. "
        f"The previous recipe was \"{old_title}\" and I'd like something different. "
        f"{context} "
        "The recipe MUST come from a real recipe website (e.g. BBC Good Food, RecipeTin Eats, Mob Kitchen, Serious Eats, etc). "
        "You MUST provide the real source_url to the specific recipe page. "
        "Return ONLY a JSON object (no markdown) with this structure: "
        '{"recipe": {"title": "...", "ingredients": ["..."], "instructions": ["..."], '
        '"total_time": "...", "servings": "...", "source_url": "https://...", "categories": ["..."]}}'
    )

    system = "You are a creative chef. Always use METRIC measurements. Every recipe must come from a real, well-known recipe website with a valid source_url. Never invent recipes. Return only valid JSON, no markdown fences."
    lang = models.get_setting(session["user_id"], "language", "en")
    if lang == "fr":
        system += (
            " The user speaks French. Recipe title, ingredients, and instructions should be in French. "
            "JSON keys must remain in English."
        )

    try:
        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=[{"role": "user", "parts": [{"text": prompt}]}],
            config={
                "system_instruction": system,
                "temperature": 0.9,
                "max_output_tokens": 2000,
            },
        )
        raw = response.text.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        parsed = json.loads(raw)
        return jsonify(parsed)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ──────────────────────────────────────
# Scrape Recipe Image from source URL
# ──────────────────────────────────────

def _scrape_image_url(url):
    """Scrape the hero image from a recipe source URL. Returns image_url or None."""
    # Gousto: use their CMS API (pages are client-rendered so normal scraping fails)
    if "gousto.co.uk" in url:
        try:
            slug = url.rstrip("/").split("/")[-1]
            api_resp = requests.get(
                f"https://production-api.gousto.co.uk/cmsreadbroker/v1/recipe/{slug}",
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=10,
            )
            if api_resp.status_code == 200:
                images = api_resp.json().get("data", {}).get("entry", {}).get("media", {}).get("images", [])
                # Pick the largest image available
                if images:
                    best = max(images, key=lambda i: i.get("width", 0))
                    return best.get("image")
        except Exception:
            pass
        return None

    try:
        resp = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/120.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "DNT": "1",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Sec-Fetch-User": "?1",
            },
            timeout=15,
            allow_redirects=True,
        )
        if resp.status_code >= 400:
            return None
        html_text = resp.text
        image_url = None

        try:
            scraper = scrape_html(html=html_text, org_url=url)
            image_url = scraper.image()
        except Exception:
            pass

        if not image_url:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_text, "html.parser")
            og = soup.find("meta", property="og:image")
            if og and og.get("content"):
                image_url = og["content"]
            if not image_url:
                tw = soup.find("meta", attrs={"name": "twitter:image"})
                if tw and tw.get("content"):
                    image_url = tw["content"]
            if not image_url:
                for img in soup.find_all("img", src=True):
                    src = img["src"]
                    if any(skip in src.lower() for skip in [".svg", "icon", "logo", "avatar", "1x1", "pixel"]):
                        continue
                    if src.startswith("data:"):
                        continue
                    image_url = src
                    break

        if image_url:
            if image_url.startswith("//"):
                image_url = "https:" + image_url
            elif image_url.startswith("/"):
                from urllib.parse import urlparse
                parsed = urlparse(url)
                image_url = f"{parsed.scheme}://{parsed.netloc}{image_url}"
            return image_url
    except Exception:
        pass
    return None


@app.route("/api/scrape-recipe-image", methods=["POST"])
@login_required_api
def api_scrape_recipe_image():
    """Scrape the hero image from a recipe source URL with multiple fallbacks."""
    data = request.get_json(force=True)
    url = data.get("source_url", "").strip()
    if not url:
        return jsonify({"error": "source_url is required"}), 400
    image_url = _scrape_image_url(url)
    if image_url:
        return jsonify({"image_url": image_url})
    return jsonify({"error": "No image found on page"}), 404


# ──────────────────────────────────────
# AI Chatbot
# ──────────────────────────────────────

SYSTEM_PROMPT = """You are a friendly and creative meal-planning assistant.
When the user asks you to suggest meals, generate a structured JSON response.
Always use METRIC measurements (grams, kilograms, millilitres, litres, Celsius) — never use cups, ounces, pounds, or Fahrenheit.

CRITICAL RULE: Every recipe you suggest MUST come from a real, existing recipe on a well-known recipe website. You MUST provide the actual source_url to the original recipe page. Never invent recipes — always reference real ones from sites like BBC Good Food, RecipeTin Eats, Mob Kitchen, Delicious Magazine, Taste.com.au, Budget Bytes, Serious Eats, Cookie and Kate, Simply Recipes, etc.
The source_url MUST be a real, working URL to the specific recipe page (not just the homepage). This is mandatory for every recipe.

You can respond in two modes:

1. RECIPE SUGGESTIONS — when the user asks for recipe ideas:
{
  "message": "A friendly text response",
  "recipes": [
    {
      "title": "Recipe Name",
      "ingredients": ["ingredient 1", "ingredient 2"],
      "instructions": ["Step 1", "Step 2"],
      "total_time": "30 min",
      "servings": "4 servings",
      "source_url": "https://www.bbcgoodfood.com/recipes/example",
      "categories": ["Dinner", "Pasta"]
    }
  ]
}

2. MEAL PLAN — when the user asks you to create/generate/plan a meal plan or weekly meals:
{
  "message": "A friendly text response",
  "meal_plan": {
    "entries": [
      {
        "day": "2026-04-06",
        "meal": "lunch",
        "recipe": {
          "title": "Recipe Name",
          "ingredients": ["ingredient 1"],
          "instructions": ["Step 1"],
          "total_time": "30 min",
          "servings": "4 servings",
          "source_url": "https://www.bbcgoodfood.com/recipes/example",
          "categories": ["Lunch"]
        }
      },
      {
        "day": "2026-04-06",
        "meal": "dinner",
        "recipe": { ... }
      }
    ]
  },
  "recipes": []
}

Days must be ISO date strings (YYYY-MM-DD). When the user asks for a week of meals, start from the next upcoming Monday and generate 7 consecutive days.
Meals must be: lunch, dinner.
Include both lunch and dinner for each day requested.
REMEMBER: source_url is MANDATORY for every recipe. Always provide a real, specific URL to the recipe on a well-known recipe website. Never set it to null.
If the user is just chatting and not asking for recipes or meal plans, return an empty recipes array and no meal_plan.
Be creative, consider dietary preferences, and suggest balanced meals.
Keep ingredient lists practical and instructions clear."""


@app.route("/api/chat", methods=["POST"])
@login_required_api
def api_chat():
    client = _get_gemini()
    if client is None:
        return jsonify({
            "error": "Gemini API key not configured. Add GEMINI_API_KEY to your .env file."
        }), 503

    data = request.get_json(force=True)
    user_message = data.get("message", "").strip()
    history = data.get("history", [])

    if not user_message:
        return jsonify({"error": "Message is required"}), 400

    contents = []
    for msg in history[-10:]:
        role = "model" if msg.get("role") == "assistant" else "user"
        contents.append({"role": role, "parts": [{"text": msg.get("content", "")}]})
    contents.append({"role": "user", "parts": [{"text": user_message}]})

    system = SYSTEM_PROMPT
    lang = models.get_setting(session["user_id"], "language", "en")
    if lang == "fr":
        system += (
            "\n\nIMPORTANT: The user speaks French. Your 'message' field MUST be written in French. "
            "Recipe titles, ingredients, and instructions should be in French. "
            "JSON keys must remain in English."
        )

    try:
        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=contents,
            config={
                "system_instruction": system,
                "temperature": 0.8,
                "max_output_tokens": 8000,
            },
        )
        raw = response.text.strip()

        try:
            if raw.startswith("```"):
                raw = re.sub(r"^```(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            parsed = {"message": raw, "recipes": []}

        return jsonify(parsed)

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ──────────────────────────────────────
# Settings API
# ──────────────────────────────────────

@app.route("/api/settings", methods=["GET"])
@login_required_api
def api_get_settings():
    return jsonify(models.get_settings(session["user_id"]))


@app.route("/api/settings", methods=["PUT"])
@login_required_api
def api_update_settings():
    data = request.get_json(force=True)
    user_id = session["user_id"]
    for key, value in data.items():
        models.set_setting(user_id, key, str(value))
    return jsonify(models.get_settings(user_id))


# ──────────────────────────────────────

if __name__ == "__main__":
    debug = os.getenv("FLASK_ENV", "development") != "production"
    app.run(host="0.0.0.0", debug=debug, port=5001)
